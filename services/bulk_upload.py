"""
Bulk Upload Service — CSV/Excel contact uploads with opt-in tracking
"""
import os
import csv
import io
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone
from dataclasses import dataclass

logger = logging.getLogger("bulk_upload")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


@dataclass
class ContactUploadResult:
    """Result of a bulk contact upload"""
    total_rows: int = 0
    successful: int = 0
    failed: int = 0
    skipped_duplicates: int = 0
    errors: List[Dict[str, Any]] = None
    contacts: List[Dict[str, Any]] = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []
        if self.contacts is None:
            self.contacts = []


class BulkUploadService:
    """Handle bulk contact and catalog uploads from CSV/Excel"""

    REQUIRED_CONTACT_FIELDS = ["phone_number"]
    OPTIONAL_CONTACT_FIELDS = ["name", "email", "source", "campaign_id", "ad_id", "opt_in", "opt_in_timestamp", "opt_in_source"]
    VALID_OPT_IN_VALUES = ["yes", "y", "true", "1", "explicit", "implicit"]

    def parse_contacts_file(self, file_content: bytes, filename: str, default_source: str = "bulk_upload") -> ContactUploadResult:
        """Parse CSV or Excel file and extract contact data"""
        result = ContactUploadResult()
        ext = os.path.splitext(filename)[1].lower()

        try:
            if ext in (".xlsx", ".xls"):
                rows = self._parse_excel(file_content)
            elif ext == ".csv":
                rows = self._parse_csv(file_content)
            else:
                result.errors.append({"row": 0, "error": f"Unsupported file format: {ext}. Use CSV or XLSX."})
                result.failed = 1
                return result

            if not rows:
                result.errors.append({"row": 0, "error": "File is empty or has no data rows"})
                result.failed = 1
                return result

            result.total_rows = len(rows)
            result = self._validate_and_process(rows, result, default_source)

        except Exception as e:
            logger.error(f"Failed to parse upload file: {e}")
            result.errors.append({"row": 0, "error": f"Parse error: {str(e)}"})
            result.failed = result.total_rows

        return result

    def _parse_csv(self, content: bytes) -> List[Dict[str, str]]:
        """Parse CSV content"""
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    def _parse_excel(self, content: bytes) -> List[Dict[str, str]]:
        """Parse Excel content"""
        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            raise ImportError("openpyxl or pandas is required for Excel uploads")

        if PANDAS_AVAILABLE:
            df = pd.read_excel(io.BytesIO(content))
            return df.to_dict(orient="records")
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            return rows

    def _validate_and_process(self, rows: List[Dict], result: ContactUploadResult, default_source: str) -> ContactUploadResult:
        """Validate rows and build contact list"""
        headers = list(rows[0].keys())
        missing_required = [f for f in self.REQUIRED_CONTACT_FIELDS if f not in headers]
        if missing_required:
            result.errors.append({"row": 0, "error": f"Missing required columns: {', '.join(missing_required)}"})
            result.failed = result.total_rows
            return result

        seen_numbers = set()
        for idx, row in enumerate(rows, start=2):
            phone = str(row.get("phone_number", "")).strip()
            if not phone:
                result.errors.append({"row": idx, "error": "Missing phone_number", "data": row})
                result.failed += 1
                continue

            phone = self._normalize_phone(phone)
            if not phone:
                result.errors.append({"row": idx, "error": f"Invalid phone number: {row.get('phone_number')}", "data": row})
                result.failed += 1
                continue

            if phone in seen_numbers:
                result.skipped_duplicates += 1
                continue
            seen_numbers.add(phone)

            opt_in_raw = str(row.get("opt_in", "yes")).lower().strip()
            has_opt_in = opt_in_raw in self.VALID_OPT_IN_VALUES
            opt_in_timestamp = row.get("opt_in_timestamp") or (datetime.now(timezone.utc).isoformat() if has_opt_in else None)
            opt_in_source = row.get("opt_in_source") or (default_source if has_opt_in else None)

            contact = {
                "phone_number": phone,
                "name": str(row.get("name", "")).strip() or None,
                "email": str(row.get("email", "")).strip() or None,
                "source": str(row.get("source", default_source)).strip(),
                "campaign_id": str(row.get("campaign_id", "")).strip() or None,
                "ad_id": str(row.get("ad_id", "")).strip() or None,
                "opt_in": has_opt_in,
                "opt_in_timestamp": opt_in_timestamp,
                "opt_in_source": opt_in_source,
                "tags": ["bulk_import"],
                "custom_fields": {"imported_at": datetime.now(timezone.utc).isoformat(), "row_number": idx},
            }
            result.contacts.append(contact)
            result.successful += 1

        return result

    def _normalize_phone(self, phone: str) -> Optional[str]:
        """Normalize phone number to E.164-like format"""
        digits = "".join(c for c in phone if c.isdigit())
        if len(digits) < 10:
            return None
        if len(digits) == 10:
            return "91" + digits
        return digits

    def parse_catalog_file(self, file_content: bytes, filename: str, business_id: str) -> Dict[str, Any]:
        """Parse CSV/Excel for catalog items"""
        result = {"total": 0, "items": [], "errors": []}
        ext = os.path.splitext(filename)[1].lower()

        try:
            if ext in (".xlsx", ".xls"):
                rows = self._parse_excel(file_content)
            elif ext == ".csv":
                rows = self._parse_csv(file_content)
            else:
                result["errors"].append(f"Unsupported format: {ext}")
                return result

            if not rows:
                result["errors"].append("Empty file")
                return result

            result["total"] = len(rows)
            for idx, row in enumerate(rows, start=2):
                try:
                    name = str(row.get("name", row.get("product_name", ""))).strip()
                    if not name:
                        result["errors"].append(f"Row {idx}: Missing name/product_name")
                        continue

                    item = {
                        "business_id": business_id,
                        "name": name,
                        "description": str(row.get("description", "")).strip(),
                        "category": str(row.get("category", "General")).strip(),
                        "price": float(row.get("price", row.get("rate", 0)) or 0),
                        "image_url": str(row.get("image_url", row.get("image", ""))).strip() or None,
                        "is_available": str(row.get("is_available", "true")).lower() not in ("false", "0", "no"),
                        "tags": [t.strip() for t in str(row.get("tags", "")).split(",") if t.strip()],
                        "variants": {},
                        "sort_order": idx,
                    }
                    result["items"].append(item)
                except Exception as e:
                    result["errors"].append(f"Row {idx}: {str(e)}")

        except Exception as e:
            logger.error(f"Catalog parse error: {e}")
            result["errors"].append(str(e))

        return result


bulk_upload = BulkUploadService()
