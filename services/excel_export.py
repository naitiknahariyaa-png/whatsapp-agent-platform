"""
Excel Export Service — Generate .xlsx reports for analytics, leads, contacts, appointments
"""
import os
import io
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone

logger = logging.getLogger("excel_export")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ExcelExportService:
    """Generate Excel files from analytics and CRM data"""

    def __init__(self):
        if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
            raise ImportError("openpyxl or pandas is required for Excel export")

    def _style_header(self, ws, row_idx: int = 1):
        """Apply header styling to a worksheet"""
        if not OPENPYXL_AVAILABLE:
            return
        header_fill = PatternFill(start_color="25D366", end_color="25D366", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[row_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def export_analytics_report(self, client_id: int, days: int = 30,
                                stats: Dict = None, daily_stats: List[Dict] = None,
                                events: List[Dict] = None) -> bytes:
        """Generate analytics report as XLSX bytes"""
        from analytics import analytics as analytics_engine
        if not stats:
            stats = analytics_engine.get_summary(client_id, days)
        if not daily_stats:
            daily_stats = analytics_engine.get_daily_stats(client_id, days)
        if not events:
            events = analytics_engine.get_events(client_id, days)

        if PANDAS_AVAILABLE:
            return self._export_with_pandas(stats, daily_stats, events)
        elif OPENPYXL_AVAILABLE:
            return self._export_with_openpyxl(stats, daily_stats, events)
        else:
            raise ImportError("No Excel library available")

    def _export_with_pandas(self, stats: Dict, daily_stats: List[Dict], events: List[Dict]) -> bytes:
        """Export using pandas"""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Summary sheet
            summary_df = pd.DataFrame([{
                "Metric": k.replace("_", " ").title(),
                "Value": v
            } for k, v in stats.items()])
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Daily stats sheet
            if daily_stats:
                daily_df = pd.DataFrame(daily_stats)
                daily_df.to_excel(writer, sheet_name="Daily Stats", index=False)

            # Events sheet
            if events:
                events_df = pd.DataFrame(events)
                events_df.to_excel(writer, sheet_name="Events", index=False)

        output.seek(0)
        return output.read()

    def _export_with_openpyxl(self, stats: Dict, daily_stats: List[Dict], events: List[Dict]) -> bytes:
        """Export using openpyxl"""
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Summary
        ws_summary.append(["Metric", "Value"])
        self._style_header(ws_summary)
        for k, v in stats.items():
            ws_summary.append([k.replace("_", " ").title(), v])

        # Daily stats
        if daily_stats:
            ws_daily = wb.create_sheet("Daily Stats")
            headers = list(daily_stats[0].keys()) if daily_stats else []
            ws_daily.append(headers)
            self._style_header(ws_daily)
            for row in daily_stats:
                ws_daily.append([row.get(h, "") for h in headers])

        # Events
        if events:
            ws_events = wb.create_sheet("Events")
            headers = list(events[0].keys()) if events else []
            ws_events.append(headers)
            self._style_header(ws_events)
            for row in events:
                ws_events.append([row.get(h, "") for h in headers])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_leads_report(self, leads: List[Dict]) -> bytes:
        """Export leads to Excel"""
        if not leads:
            leads = []

        if PANDAS_AVAILABLE:
            df = pd.DataFrame(leads)
            output = io.BytesIO()
            df.to_excel(output, index=False, engine="openpyxl")
            output.seek(0)
            return output.read()
        elif OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads"
            if leads:
                headers = list(leads[0].keys())
                ws.append(headers)
                self._style_header(ws)
                for lead in leads:
                    ws.append([lead.get(h, "") for h in headers])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
        else:
            raise ImportError("No Excel library available")

    def export_contacts_report(self, contacts: List[Dict]) -> bytes:
        """Export contacts to Excel"""
        if not contacts:
            contacts = []

        if PANDAS_AVAILABLE:
            df = pd.DataFrame(contacts)
            output = io.BytesIO()
            df.to_excel(output, index=False, engine="openpyxl")
            output.seek(0)
            return output.read()
        elif OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contacts"
            if contacts:
                headers = list(contacts[0].keys())
                ws.append(headers)
                self._style_header(ws)
                for contact in contacts:
                    ws.append([contact.get(h, "") for h in headers])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
        else:
            raise ImportError("No Excel library available")


excel_export = ExcelExportService()
